import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
from openpyxl import Workbook


def process_and_save(vcf_paths, save_path, progress_var, btn_sel, btn_run, pct_threshold):
    try:
        # 1) VCF dosyalarından veriyi oku
        variant_info = {}  # variantKey -> (chrom, pos, ref, alt, gene, protein, class)
        sample_data = {}   # sample_id -> { variantKey -> (dp, ad, pct) }
        sample_ids = []
        total_vcf = len(vcf_paths)

        for idx, path in enumerate(vcf_paths):
            sid = os.path.basename(path).split('_')[0]
            sample_ids.append(sid)
            sample_data[sid] = {}
            with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                for line in f:
                    if line.startswith('#'): continue
                    cols = line.strip().split('\t')
                    if len(cols) < 10: continue
                    chrom, pos, _, ref, alt, *_ , info_str, fmt_str, sample_str = cols[:10]
                    vk = f"{chrom}_{pos}_{ref}>{alt}"

                    # Statik anotasyonları kaydet
                    if vk not in variant_info:
                        info = {kv.split('=')[0]: kv.split('=')[1]
                                for kv in info_str.split(';') if '=' in kv}
                        variant_info[vk] = (
                            chrom,
                            int(pos),
                            ref,
                            alt,
                            info.get('GENE_SYMBOL',''),
                            info.get('HGVS_PROTEIN',''),
                            info.get('ING_CLASSIFICATION','')
                        )

                    # Depth ve allele count
                    keys = fmt_str.split(':')
                    vals = sample_str.split(':')
                    fmt = dict(zip(keys, vals))
                    dp = int(fmt.get('DP', '0')) if fmt.get('DP', '').isdigit() else 0
                    ad_vals = fmt.get('AD', '0,0').split(',')
                    ad = int(ad_vals[1]) if len(ad_vals) > 1 and ad_vals[1].isdigit() else 0
                    pct = round(ad / dp * 100, 2) if dp > 0 else 0.0
                    sample_data[sid][vk] = (dp, ad, pct)

            # İlerleme: VCF okuma %0–30
            progress_var.set(int((idx + 1) / total_vcf * 30))

        # 2) Filter variant_keys by threshold
        variant_keys = sorted(k for k in variant_info if
                              any(sample_data[s].get(k, (0,0,0.0))[2] >= pct_threshold
                                  for s in sample_ids))
        nvar = len(variant_keys)

        # 3) Excel Workbook ve sayfalar
        wb = Workbook()
        ws_readme = wb.active
        ws_readme.title = 'README'
        ws_all = wb.create_sheet('AllVariants')
        ws_common = wb.create_sheet('CommonVariants')
        ws_non = wb.create_sheet('NonCommonVariants')
        ws_tek = wb.create_sheet('AllVariants_tek_sutun')

        # README içeriği
        ws_readme.append(["NGS Varyant Analiz Aracı"])
        ws_readme.append([f"Yüzde eşik değeri: {pct_threshold}%"])
        ws_readme.append(["Sayfalar:"])
        for name in ['AllVariants','CommonVariants','NonCommonVariants','AllVariants_tek_sutun']:
            ws_readme.append([f"- {name}"])

        # 4) Başlıklar
        static_h = ['VariantKey','Chromosome','Position','Ref','Alt',
                    'Gene','ProteinChange','Classification']
        dp_h = [f'DP_{s}' for s in sample_ids]
        ad_h = [f'AD_{s}' for s in sample_ids]
        pct_h = [f'Pct_{s}' for s in sample_ids]
        comb_h = [f'Comb_{s}' for s in sample_ids]
        hdr_all = static_h + dp_h + ad_h + pct_h + ['OccurrenceCount','OccurrenceType']
        hdr_tek = static_h + comb_h + ['OccurrenceCount','OccurrenceType']

        for ws in (ws_all, ws_common, ws_non):
            ws.append(hdr_all)
        ws_tek.append(hdr_tek)

        # 5) Varyantları satır satır yaz
        for idx, vk in enumerate(variant_keys):
            chrom, pos, ref, alt, gene, prot, clas = variant_info[vk]
            occ_count = 0
            row_dp, row_ad, row_pct, row_comb = [], [], [], []
            for s in sample_ids:
                dp, ad, pct = sample_data[s].get(vk, (0,0,0.0))
                if pct >= pct_threshold:
                    occ_count += 1
                row_dp.append(dp if dp>0 else "")
                row_ad.append(ad if ad>0 else "")
                row_pct.append(pct if pct>=pct_threshold else "")
                row_comb.append(f"(%{int(pct)}) {ad}/{dp}" if pct>=pct_threshold else "")

            occ_type = 'Common' if occ_count == len(sample_ids) else 'NonCommon'
            base = [vk, chrom, pos, ref, alt, gene, prot, clas]
            all_row = base + row_dp + row_ad + row_pct + [occ_count, occ_type]
            tek_row = base + row_comb + [occ_count, occ_type]

            ws_all.append(all_row)
            ws_tek.append(tek_row)
            if occ_type == 'Common':
                ws_common.append(all_row)
            else:
                ws_non.append(all_row)

            # İlerleme: Excel yazma %30–100
            if nvar>0:
                progress_var.set(30 + int(idx / nvar * 70))

        # 6) Kaydet
        wb.save(save_path)
        progress_var.set(100)
        messagebox.showinfo("Bitti", f"Excel oluşturuldu:\n{save_path}")

    except Exception as e:
        messagebox.showerror("Hata", str(e))
    finally:
        btn_sel.config(state=tk.NORMAL)
        btn_run.config(state=tk.NORMAL)

# GUI fonksiyonları

def on_select():
    paths = filedialog.askopenfilenames(
        title="VCF Dosyalarını Seçin", filetypes=[("VCF","*.vcf"),("Tüm","*.*")]
    )
    if paths:
        listbox.delete(0, tk.END)
        for p in paths:
            listbox.insert(tk.END, p)
        btn_run.config(state=tk.NORMAL)

def on_run():
    vcfs = listbox.get(0, tk.END)
    if not vcfs:
        messagebox.showwarning("Uyarı", "Önce VCF dosyalarını seçin.")
        return
    pct_threshold = simpledialog.askfloat(
        "Eşik Değeri", "Değerlendirilecek minimum yüzde (0-100):",
        minvalue=0, maxvalue=100, initialvalue=0)
    if pct_threshold is None:
        return
    savep = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel","*.xlsx")],
        title="Sonucu Kaydet"
    )
    if not savep:
        return
    btn_sel.config(state=tk.DISABLED)
    btn_run.config(state=tk.DISABLED)
    progress_var.set(0)
    threading.Thread(
        target=process_and_save,
        args=(vcfs, savep, progress_var, btn_sel, btn_run, pct_threshold),
        daemon=True
    ).start()

# --- TKinter Arayüzü ---
root = tk.Tk()
root.title("NGS Varyant Analiz Aracı")
frm = ttk.Frame(root, padding=10)
frm.grid(sticky="nsew")

# Açıklama Label
ttk.Label(frm, text=(
    "1) VCF Dosyalarını Seçin.\n"
    "2) Minimum yüzde eşik değerini girin.\n"
    "3) Excel Olarak Kaydet ve İşle\n"
    "4) İlerlemeyi izleyin."
)).grid(row=0, column=0, columnspan=2, pady=(0,10))

# Dosya Listbox
listbox = tk.Listbox(frm, width=80, height=6)
listbox.grid(row=1, column=0, columnspan=2, pady=(0,10))

# Seç Butonu
btn_sel = ttk.Button(frm, text="VCF Dosyalarını Seç", command=on_select)
btn_sel.grid(row=2, column=0, sticky="ew", padx=(0,5))

# Kaydet ve Çalıştır Butonu
btn_run = ttk.Button(frm, text="Excel Olarak Kaydet ve İşle", command=on_run, state=tk.DISABLED)
btn_run.grid(row=2, column=1, sticky="ew", padx=(5,0))

# İlerleme Çubuğu
progress_var = tk.IntVar()
progress = ttk.Progressbar(frm, maximum=100, variable=progress_var)
progress.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(10,0))

# Pencere Ayarları
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)
root.mainloop()
